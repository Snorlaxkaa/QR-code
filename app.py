from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import mysql.connector
from db import get_connection
from openpyxl import Workbook
from datetime import datetime
import os
import io
import qrcode
from qrcode.constants import ERROR_CORRECT_Q
import base64
from datetime import datetime
from flask import send_file
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from user import User

app = Flask(__name__)
app.secret_key = "super_secret_key_123"

# 初始化 Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'  # 設定登入頁面的路由
login_manager.login_message = '請先登入以繼續'

@login_manager.user_loader
def load_user(user_id):
    return User.get(user_id)

# ----------- 登入頁面 -----------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.get_by_username(username)
        
        if user and user.password == password:
            login_user(user)
            flash('登入成功！', 'success')
            return redirect(url_for('admin_panel'))
        else:
            flash('帳號或密碼錯誤！', 'error')
    
    return render_template('login.html')

# ----------- 登出 -----------
@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('已登出！', 'success')
    return redirect(url_for('login'))

# ----------- 後台主頁 -----------
@app.route('/', methods=['GET', 'POST'])
@app.route('/index', methods=['GET', 'POST'])
@login_required
def admin_panel():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    # --- 基本參數 ---
    nid = request.args.get('nid', '').strip()
    date_start = request.args.get('date_start')
    date_end = request.args.get('date_end')
    page = request.args.get('page', 1, type=int)
    per_page = 30

    print(f"搜尋參數：nid={nid}, date_start={date_start}, date_end={date_end}")  # 除錯用

    # --- 驗證日期格式 ---
    for var_name, var_value in [('開始日期', date_start), ('結束日期', date_end)]:
        if var_value:
            try:
                datetime.strptime(var_value, '%Y-%m-%d')
            except ValueError:
                flash(f'{var_name}格式不正確', 'error')
                if var_name == '開始日期': date_start = None
                else: date_end = None

    # 組合查詢
    base_query = """
        SELECT serial_no, name, id_number, service_start, service_end,
               service_item, service_content, service_hours, service_minutes,
               served_people_count, transport_fee, meal_fee, service_area,
               remarks, import_action, foreign_service_count, domestic_service_count
        FROM service_records
        WHERE 1=1
    """
    query = base_query
    params = []

    if nid:
        query += " AND LOWER(id_number) = LOWER(%s)"
        params.append(nid)
    else:
        # 如果沒有指定身分證號，則套用日期條件
        if date_start and date_end:
            query += " AND DATE(service_start) BETWEEN %s AND %s"
            params.extend([date_start, date_end])
        elif date_start:
            query += " AND DATE(service_start) >= %s"
            params.append(date_start)
        elif date_end:
            query += " AND DATE(service_start) <= %s"
            params.append(date_end)
        else:
            # 預設顯示當月資料
            query += " AND YEAR(service_start)=YEAR(CURDATE()) AND MONTH(service_start)=MONTH(CURDATE())"

    query += " ORDER BY service_start DESC"

    # --- 執行查詢 ---
    cursor.execute(query, tuple(params))
    all_records = cursor.fetchall()

    # --- 分頁 ---
    total_records = len(all_records)
    total_pages = max(1, (total_records + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    current_page_records = all_records[start_idx:end_idx]

    # --- 計算總工時 ---
    total_hours = sum(r.get('service_hours') or 0 for r in all_records)
    total_minutes = sum(r.get('service_minutes') or 0 for r in all_records)
    total_hours += total_minutes // 60
    total_minutes = total_minutes % 60

    # --- 下方完整資料表分頁 ---
    # 獲取完整記錄的頁碼
    all_records_page = request.args.get('all_records_page', 1, type=int)
    
    # 獲取所有記錄
    cursor.execute("SELECT * FROM service_records ORDER BY serial_no DESC")
    all_records_data = cursor.fetchall()
    
    # 計算完整記錄的分頁
    total_all_records = len(all_records_data)
    total_all_pages = max(1, (total_all_records + per_page - 1) // per_page)
    all_records_page = max(1, min(all_records_page, total_all_pages))
    
    # 對完整記錄進行分頁
    start_idx_all = (all_records_page - 1) * per_page
    end_idx_all = start_idx_all + per_page
    records = all_records_data[start_idx_all:end_idx_all]
    
    cursor.close()
    conn.close()

    return render_template(
        'index.html',
        records=records,
        user=current_user.username,  # 使用登入用戶的名稱
        nid=nid,
        personal_records=current_page_records,
        total_hours=total_hours,
        total_minutes=total_minutes,
        date_start=date_start,
        date_end=date_end,
        page=page,
        total_pages=total_pages,
        total_records=total_records,
        all_records_page=all_records_page,
        total_all_pages=total_all_pages,
        total_all_records=total_all_records
    )

# ----------- 匯出資料（改為匯出搜尋結果）-----------
@app.route('/export_xlsx', methods=['GET', 'POST'])
def export_xlsx():
    nid = request.values.get('nid')
    date_start = request.values.get('date_start')
    date_end = request.values.get('date_end')

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    # ✅ 基本查詢
    query = """
        SELECT name, id_number, service_start, service_end,
               service_item, service_content, service_hours, service_minutes,
               served_people_count, transport_fee, meal_fee, service_area,
               remarks, import_action, foreign_service_count, domestic_service_count
        FROM service_records
        WHERE 1=1
    """
    params = []

    # ✅ 條件（依據輸入）
    if nid:
        query += " AND LOWER(id_number) = LOWER(%s)"
        params.append(nid)

    if date_start and date_end:
        query += " AND DATE(service_start) BETWEEN %s AND %s"
        params.extend([date_start, date_end])
    elif date_start:
        query += " AND DATE(service_start) >= %s"
        params.append(date_start)
    elif date_end:
        query += " AND DATE(service_start) <= %s"
        params.append(date_end)

    query += " ORDER BY service_start DESC"

    cursor.execute(query, tuple(params))
    records = cursor.fetchall()
    cursor.close()
    conn.close()

    # ✅ 生成 Excel
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Search Results"

    if records:
        header_map = {
            "name": "姓名",
            "id_number": "身分證",
            "service_start": "上班時間",
            "service_end": "下班時間",
            "service_item": "服務項目",
            "service_content": "服務內容",
            "service_hours": "工時",
            "service_minutes": "分鐘",
            "served_people_count": "服務人數",
            "transport_fee": "交通費",
            "meal_fee": "餐費",
            "service_area": "服務地區",
            "remarks": "備註",
            "import_action": "匯入動作",
            "foreign_service_count": "海外服務",
            "domestic_service_count": "國內服務"
        }

        keys = [k for k in records[0].keys() if k in header_map]
        ws.append([header_map[k] for k in keys])

        # 設定標題列置中對齊
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for r in records:
            for k, v in r.items():
                if v is None:
                    r[k] = ""
            row = ws.append([r[k] for k in keys])
            # 設定資料列置中對齊
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # ✅ 自動欄寬（優化版）
        from datetime import datetime

        def get_column_width(cell_value):
            if cell_value is None:
                return 0
            if isinstance(cell_value, datetime):
                return 20  # 日期時間固定寬度
            
            str_value = str(cell_value)
            # 計算中文字元的數量
            chinese_count = sum(1 for char in str_value if '\u4e00' <= char <= '\u9fff')
            # 其他字元的數量
            other_count = len(str_value) - chinese_count
            # 中文字元寬度為 2.1，其他字元為 1
            return chinese_count * 2.1 + other_count

        for col in ws.columns:
            max_length = max(get_column_width(cell.value) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # ✅ 儲存與回傳
  
    
    # 設定檔名
    if date_start or date_end:
        # 有日期範圍的情況
        if date_start and date_end:
            filename = f"工時紀錄_{date_start}至{date_end}"
        elif date_start:
            filename = f"工時紀錄_{date_start}起"
        elif date_end:
            filename = f"工時紀錄_至{date_end}"
            
        # 如果有身分證號，加到檔名前面
        if nid:
            filename = f"{nid}_{filename}"
        
        filename += ".xlsx"
    else:
        # 只有身分證號或完全沒有搜尋條件
        if nid:
            filename = f"{nid}.xlsx"
        else:
            filename = f"全部資料_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    file_path = os.path.join(os.getcwd(), filename)
    wb.save(file_path)
    return send_file(file_path, as_attachment=True)

# ----------- 編輯資料 -----------
@app.route('/edit/<int:serial_no>', methods=['GET', 'POST'])
def edit(serial_no):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        fields = [
            'name', 'id_number', 'service_start', 'service_end',
            'service_item', 'service_content', 'service_hours', 'service_minutes',
            'served_people_count', 'transport_fee', 'meal_fee',
            'service_area', 'remarks', 'import_action',
            'foreign_service_count', 'domestic_service_count'
        ]

        values = []
        for f in fields:
            v = request.form.get(f)
            
            # 處理日期時間格式
            if f in ['service_start', 'service_end'] and v:
                try:
                    # 將HTML datetime-local格式轉換為MySQL datetime格式
                    v = datetime.strptime(v, '%Y-%m-%dT%H:%M').strftime('%Y-%m-%d %H:%M:00')
                except ValueError:
                    v = None
            
            # 處理數值欄位
            elif f in ["service_hours", "service_minutes", "served_people_count",
                      "transport_fee", "meal_fee",
                      "foreign_service_count", "domestic_service_count"]:
                try:
                    v = int(v) if v and v.strip() else 0
                except ValueError:
                    v = 0
            
            # 處理其他可為空的欄位
            elif v is None or v.strip() == "":
                v = None
                
            values.append(v)

        # ✅ 最後加上 serial_no 讓 SQL 的 WHERE 能對應上
        values.append(serial_no)

        sql = """
            UPDATE service_records
            SET name=%s, id_number=%s, service_start=%s, service_end=%s,
                service_item=%s, service_content=%s, service_hours=%s, service_minutes=%s,
                served_people_count=%s, transport_fee=%s, meal_fee=%s,
                service_area=%s, remarks=%s, import_action=%s,
                foreign_service_count=%s, domestic_service_count=%s
            WHERE serial_no=%s
        """
        try:
            cursor.execute(sql, tuple(values))
            conn.commit()
            flash("✅ 資料已更新", "success")
        except Exception as e:
            flash(f"❌ 錯誤：{e}", "error")
        finally:
            cursor.close()
            conn.close()
        return redirect(url_for('admin_panel'))

    # 讀取指定編號資料
    cursor.execute("SELECT * FROM service_records WHERE serial_no = %s", (serial_no,))
    record = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('edit.html', record=record)


# ----------- 刪除資料 -----------
@app.route('/delete/<int:serial_no>')
def delete(serial_no):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM service_records WHERE serial_no = %s", (serial_no,))
    conn.commit()
    cursor.close()
    conn.close()
    flash('資料已刪除', 'info')
    return redirect(url_for('admin_panel'))

# ----------- 生成 QRcode -----------
@app.route("/qrcode", methods=["GET", "POST"])
def qrcode_page():
    qr_base64 = None
    user_name = None

    if request.method == "POST":
        name = request.form.get("name")
        id_number = request.form.get("id_number")

        if not name or not id_number:
            flash("請完整輸入姓名與身分證字號！")
        else:
            user_name = name
            qr_text = f"姓名：{name}\n身分證字號：{id_number}"
            qr = qrcode.QRCode(
                version=1,
                error_correction=ERROR_CORRECT_Q,
                box_size=10,
                border=4
            )
            qr.add_data(qr_text)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")

            buffer = io.BytesIO()
            img.save(buffer, format="PNG")
            buffer.seek(0)
            qr_base64 = base64.b64encode(buffer.read()).decode("utf-8")

    return render_template("QRcode.html", qr_base64=qr_base64, user_name=user_name)

if __name__ == '__main__':
    app.run(debug=True)
