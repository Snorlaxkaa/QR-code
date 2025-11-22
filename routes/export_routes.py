"""
Excel 匯出相關功能 - 支援多種格式
"""
from flask import request, make_response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from datetime import datetime
import io
from db import get_connection
from flask_login import login_required
from collections import defaultdict


def register_export_routes(app):
    """註冊匯出相關的路由"""
    
    # ----------- 匯出資料（支援多格式）-----------
    @app.route('/export_xlsx', methods=['GET', 'POST'])
    @login_required
    def export_xlsx():
        nid = request.values.get('nid', '').strip()
        date_start = request.values.get('date_start')
        date_end = request.values.get('date_end')
        export_format = request.values.get('format', 'detailed')  # 預設詳細格式

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # ✅ 基本查詢 - 使用 serial_number
        query = """
            SELECT name, id_number, service_start, service_end,
                service_item, service_content, service_hours, service_minutes,
                served_people_count, transport_fee, meal_fee, service_area,
                remarks, import_action, serial_number, foreign_service_count, domestic_service_count
            FROM service_records
            WHERE 1=1
        """
        params = []

        # ✅ 條件（依據輸入）
        if nid:
            query += " AND LOWER(id_number) = LOWER(%s)"
            params.append(nid)

        if date_start and date_end:
            query += " AND DATE(service_start) BETWEEN %s AND %s"
            params.extend([date_start, date_end])
        elif date_start:
            query += " AND DATE(service_start) >= %s"
            params.append(date_start)
        elif date_end:
            query += " AND DATE(service_start) <= %s"
            params.append(date_end)

        query += " ORDER BY service_start DESC"

        cursor.execute(query, tuple(params))
        records = cursor.fetchall()
        cursor.close()
        conn.close()

        # ✅ 根據格式選擇不同的匯出方式
        if export_format == 'summary':
            return export_summary_format(records, nid, date_start, date_end)
        else:  # detailed
            return export_detailed_format(records, nid, date_start, date_end)


def export_detailed_format(records, nid, date_start, date_end):
    """詳細格式（逐筆記錄）"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "搜尋結果"

    header_map = {
        "name": "姓名",
        "id_number": "身分證",
        "service_start": "上班時間",
        "service_end": "下班時間",
        "service_item": "服務項目",
        "service_content": "服務內容",
        "service_hours": "小時",
        "service_minutes": "分鐘",
        "served_people_count": "受服務人次",
        "transport_fee": "交通費",
        "meal_fee": "誤餐費",
        "service_area": "服務地區",
        "remarks": "備註",
        "import_action": "匯入動作",
        "serial_number": "序號",  # ✅ 改成 serial_number
        "foreign_service_count": "國外參與服務人次",
        "domestic_service_count": "國內參與服務人次"
    }

    if records:
        keys = [k for k in records[0].keys() if k in header_map]
        ws.append([header_map[k] for k in keys])

        # 設定表頭樣式
        header_fill = PatternFill(start_color="D0E8FF", end_color="D0E8FF", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 新增資料
        for r in records:
            for k, v in r.items():
                if v is None:
                    r[k] = ""
                # 特別處理 serial_number - 如果是 None 就用空字串
                if k == 'serial_number' and r[k] == "":
                    r[k] = ""
            row = ws.append([r[k] for k in keys])
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        # 自動欄寬
        def get_column_width(cell_value):
            if cell_value is None:
                return 0
            if isinstance(cell_value, datetime):
                return 20
            
            str_value = str(cell_value)
            chinese_count = sum(1 for char in str_value if '\u4e00' <= char <= '\u9fff')
            other_count = len(str_value) - chinese_count
            return chinese_count * 2.1 + other_count

        for col in ws.columns:
            max_length = max(get_column_width(cell.value) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    else:
        ws.append(["無資料"])

    # 設定檔名（使用查詢參數，避免編碼問題）
    filename_parts = []
    
    if nid:
        filename_parts.append(nid)
    
    if date_start and date_end:
        filename_parts.append(f"{date_start}_to_{date_end}")
    elif date_start:
        filename_parts.append(f"from_{date_start}")
    elif date_end:
        filename_parts.append(f"to_{date_end}")
    
    filename_parts.append(f"detailed_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    filename = "_".join(filename_parts) + ".xlsx"

    # 回傳
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    
    # 使用簡單的檔名（只用 ASCII 字符和日期）
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response


def export_summary_format(records, nid, date_start, date_end):
    """統計格式（按人員統計）"""
    
    # 按姓名分組並統計
    grouped_data = defaultdict(lambda: {
        'id_number': '',
        'service_dates': [],
        'service_items': set(),
        'service_contents': set(),
        'total_hours': 0,
        'total_minutes': 0,
        'total_people': 0,
        'total_transport': 0,
        'total_meal': 0,
        'service_areas': set(),
        'remarks': [],
        'import_actions': set(),
        'serials': [],
        'foreign_count': 0,
        'domestic_count': 0
    })

    # 統計數據
    for r in records:
        name = r['name']
        grouped_data[name]['id_number'] = r['id_number']
        
        if r['service_start']:
            service_date = r['service_start'].strftime('%Y-%m-%d')
            grouped_data[name]['service_dates'].append(service_date)
        
        if r['service_item']:
            grouped_data[name]['service_items'].add(str(r['service_item']))
        if r['service_content']:
            grouped_data[name]['service_contents'].add(str(r['service_content']))
        
        grouped_data[name]['total_hours'] += (r['service_hours'] or 0)
        grouped_data[name]['total_minutes'] += (r['service_minutes'] or 0)
        grouped_data[name]['total_people'] += (r['served_people_count'] or 0)
        grouped_data[name]['total_transport'] += (r['transport_fee'] or 0)
        grouped_data[name]['total_meal'] += (r['meal_fee'] or 0)
        
        if r['service_area']:
            grouped_data[name]['service_areas'].add(str(r['service_area']))
        
        if r['remarks']:
            grouped_data[name]['remarks'].append(str(r['remarks']))
        
        if r['import_action']:
            grouped_data[name]['import_actions'].add(str(r['import_action']))
        
        # ✅ 修正：使用 serial_number 而不是 serial_no，並處理 None 值
        serial_num = r['serial_number'] if r['serial_number'] else ""
        if serial_num:  # 只在有值時才加入
            grouped_data[name]['serials'].append(str(serial_num))
        grouped_data[name]['foreign_count'] += (r['foreign_service_count'] or 0)
        grouped_data[name]['domestic_count'] += (r['domestic_service_count'] or 0)

    # 處理分鐘進位
    for name in grouped_data:
        data = grouped_data[name]
        if data['total_minutes'] >= 60:
            extra_hours = data['total_minutes'] // 60
            data['total_hours'] += extra_hours
            data['total_minutes'] = data['total_minutes'] % 60

    # 生成 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "服務統計"

    headers = [
        "姓名", "身分證字號", "服務日期起", "服務日期迄",
        "服務項目", "服務內容", "服務時數-小時", "服務時數-分鐘",
        "受服務人次", "交通費", "誤餐費", "服務區域",
        "備註", "匯入動作", "序號", "國外參與服務人次", "國內參與服務人次"
    ]

    ws.append(headers)

    # 設定表頭樣式
    header_fill = PatternFill(start_color="D0E8FF", end_color="D0E8FF", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # 新增數據
    for name in sorted(grouped_data.keys()):
        data = grouped_data[name]
        
        dates = sorted(data['service_dates'])
        date_start_str = dates[0] if dates else ""
        date_end_str = dates[-1] if dates else ""
        
        row = [
            name,
            data['id_number'],
            date_start_str,
            date_end_str,
            '、'.join(sorted(data['service_items'])) if data['service_items'] else "",
            '、'.join(sorted(data['service_contents'])) if data['service_contents'] else "",
            data['total_hours'],
            data['total_minutes'],
            data['total_people'],
            data['total_transport'],
            data['total_meal'],
            '、'.join(sorted(data['service_areas'])) if data['service_areas'] else "",
            '；'.join(data['remarks']) if data['remarks'] else "",
            '、'.join(sorted(data['import_actions'])) if data['import_actions'] else "",
            '、'.join(data['serials']),
            data['foreign_count'],
            data['domestic_count']
        ]
        
        ws.append(row)
        
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    # 自動欄寬
    column_widths = [12, 14, 13, 13, 12, 12, 12, 12, 10, 10, 10, 12, 15, 12, 12, 14, 14]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = 'A2'

    # 設定檔名（使用查詢參數，避免編碼問題）
    filename_parts = []
    
    if nid:
        filename_parts.append(nid)
    
    if date_start and date_end:
        filename_parts.append(f"{date_start}_to_{date_end}")
    elif date_start:
        filename_parts.append(f"from_{date_start}")
    elif date_end:
        filename_parts.append(f"to_{date_end}")
    
    filename_parts.append(f"summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    filename = "_".join(filename_parts) + ".xlsx"

    # 回傳
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    
    # 使用 RFC 5987 正確編碼中文檔名
    try:
        # 嘗試用 UTF-8 編碼
        filename_encoded = filename.encode('utf-8').decode('utf-8')
        response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{filename_encoded}"
    except:
        # 如果失敗，使用簡化的檔名
        response.headers['Content-Disposition'] = f"attachment; filename=export.xlsx"
    
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response